import openpyxl
import logging
import sys
addr=(str(sys.argv[1])).replace("\\","\\"+"\\")
newaddr=addr.replace(addr[:rfind(".")],addr[:rfind(".")]+"_new")
logging.basicConfig(filename='manipulation_log.log',level=logging.DEBUG,format='%(asctime)s %(levelname)s %(message)s')
def updatedString(string):
    ListofSubstring=string.split("\n")
    for index, substring in enumerate(ListofSubstring):

        if(substring.count(".")==2):
            substring = substring.replace(substring[substring.rfind("."):substring.find("=")], substring[substring.rfind("."):substring.find("=")].upper())
            ListofSubstring[index]=substring
        elif(substring.count(".")==4):
            if "and" in substring:
                ListofSubSubString=substring.split("and")
                for index1,SubSubstring in enumerate(ListofSubSubString):
                    SubSubstring = SubSubstring.replace(SubSubstring[SubSubstring.rfind("."):SubSubstring.find("=")],SubSubstring[SubSubstring.rfind("."):SubSubstring.find("=")].upper())
                    ListofSubSubString[index1]=SubSubstring
                substringupdate=' and '.join(ListofSubSubString)
            if "or" in substring:
                ListofSubSubString=substring.split("or")
                for index1,SubSubstring in enumerate(ListofSubSubString):
                    SubSubstring = SubSubstring.replace(SubSubstring[SubSubstring.rfind("."):SubSubstring.find("=")],SubSubstring[SubSubstring.rfind("."):SubSubstring.find("=")].upper())
                    ListofSubSubString[index1]=SubSubstring
                substringupdate=' or '.join(ListofSubSubString)
            ListofSubstring[index]=substringupdate
    newstr='\n'.join(ListofSubstring)
    return newstr
logging.info('Opening the File ')
wb=openpyxl.load_workbook(addr)
logging.info('File opened')
ListofSheet=wb.get_sheet_names()
for sheetname in ListofSheet:
    logging.debug('Opened Sheet %s ', sheetname)
    ActiveSheet=wb.get_sheet_by_name(sheetname)
    if (sheetname == 'Parameters'):
        for i in range(1,ActiveSheet.max_column+1):
            if ActiveSheet.cell(row=1,column=i).value == 'Parameter Name':
                for j in range(2,ActiveSheet.max_row+1):
                    if ActiveSheet.cell(row=j,column=i).value == None:
                        continue
                    else:
                        ActiveSheet.cell(row=j, column=i).value = str(ActiveSheet.cell(row=j,column=i).value).upper()
            elif (ActiveSheet.cell(row=1,column=i).value == "SON Audit Rule" or "SON Enforcement Rule"):
                for j in range(2, ActiveSheet.max_row + 1):
                    if str(ActiveSheet.cell(row=j,column=i).value).startswith("if(RNC") or str(ActiveSheet.cell(row=j,column=i).value).startswith("expr(RNC"):
                        ActiveSheet.cell(row=j,column=i).value=updatedString(str(ActiveSheet.cell(row=j,column=i).value))
    else:
        for i in range(1,ActiveSheet.max_column+1):
            if (str(ActiveSheet.cell(row=1,column=i).value).startswith("ID")):
                str1= str(ActiveSheet.cell(row=1,column=i).value)
                ActiveSheet.cell(row=1,column=i).value = str1.replace(str1[str1.rfind("."):], str1[str1.rfind("."):].upper())
                break
        for i in range(1,ActiveSheet.max_column+1):
            if ActiveSheet.cell(row=2, column=i).value == 'Parameter Name':
                for j in range(3,ActiveSheet.max_row+1):
                    if ActiveSheet.cell(row=j,column=i).value == None:
                        continue
                    else:
                        ActiveSheet.cell(row=j, column=i).value = str(ActiveSheet.cell(row=j,column=i).value).upper()
    logging.debug('Manipulated sheet %s successfully ', sheetname)
logging.debug("Saving File")
wb.save(newaddr)
logging.debug('File Saved and Successfully Completed job')

